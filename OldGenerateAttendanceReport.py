from tkinter import filedialog
import openpyxl
from openpyxl.styles import PatternFill
import os
import datetime, timedelta
from openpyxl.styles.alignment import Alignment
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side
import tkinter


class student:
    def __init__(self, id, name, actualHours, scheduledHours, attendancePercentage):
        self.id = id
        self.name = name
        self.scheduledHours = scheduledHours
        self.actualHours = actualHours
        self.attendancePercentage = attendancePercentage
        self.fieldTrip = ""
        self.studentType = ""
        self.hourAmount = ""
        self.endDate = ""
        self.LOA = ""
        self.notes = ""
        self.VA = False
    
    
    
def createTemplate():
    thin_border = Border(left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin'))
        
    sheet.column_dimensions['A'].width = 3
    sheet.column_dimensions['B'].width = 5.33
    sheet.column_dimensions['C'].width = 25.67
    sheet.column_dimensions['D'].width = 12
    sheet.column_dimensions['E'].width = 10.5
    sheet.column_dimensions['F'].width = 10.33
    sheet.column_dimensions['G'].width = 6.67
    sheet.column_dimensions['H'].width = 10
    sheet.column_dimensions['I'].width = 27
    fontStyle = Font(bold=True, size=11)
    titleFontStyle = Font(bold=True, size=24)
    cell = "B1"
    time = datetime.datetime.now().strftime('%Y-%m-%d')
    sheet[cell].value = "STUDENT HOURS AS OF " + time
    sheet[cell].font = titleFontStyle
    sheet[cell].border = thin_border
    cell = "B2"
    sheet[cell].value = "Key"
    sheet[cell].font = fontStyle
    sheet[cell].border = thin_border
    sheet[cell].alignment = Alignment(horizontal="center", vertical="center")
    sheet.merge_cells('B2:B3')
    cell = "C2"
    sheet[cell].value = "Name"
    sheet[cell].font = fontStyle
    sheet[cell].border = thin_border
    sheet[cell].alignment = Alignment(horizontal="center", vertical="center")
    sheet.merge_cells('C2:C3')
    cell = "D2"
    sheet[cell].value = "Class"
    sheet[cell].font = fontStyle
    sheet[cell].border = thin_border
    sheet[cell].alignment = Alignment(horizontal="center", vertical="center")
    sheet.merge_cells('D2:D3')
    cell = "E2"
    sheet[cell].value = "Hours"
    sheet[cell].font = fontStyle
    sheet[cell].border = thin_border
    sheet[cell].alignment = Alignment(horizontal="center", vertical="center")
    sheet.merge_cells('E2:G2')
    cell = "H2"
    sheet[cell].value = "End Date"
    sheet[cell].font = fontStyle
    sheet[cell].border = thin_border
    sheet[cell].alignment = Alignment(horizontal="center", vertical="center")
    sheet.merge_cells('H2:H3')
    cell = "I2"
    sheet[cell].value = "Note"
    sheet[cell].font = fontStyle
    sheet[cell].border = thin_border
    sheet[cell].alignment = Alignment(horizontal="center", vertical="center")
    sheet.merge_cells('I2:I3')
    cell = "E3"
    sheet[cell].value = "Scheduled"
    sheet[cell].font = fontStyle
    sheet[cell].border = thin_border
    sheet[cell].alignment = Alignment(horizontal="center", vertical="center")
    cell = "F3"
    sheet[cell].value = "Actual"
    sheet[cell].font = fontStyle
    sheet[cell].border = thin_border
    sheet[cell].alignment = Alignment(horizontal="center", vertical="center")
    cell = "G3"
    sheet[cell].value = "%"
    sheet[cell].font = fontStyle
    sheet[cell].border = thin_border
    sheet[cell].alignment = Alignment(horizontal="center")
    
def fillReport(students, givenRow):
    
    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
    
    
    allStudents = students
    row = givenRow
    for x in range(len(allStudents)):
        numbercell = 'A' + str(row)
        idcell = 'B' + str(row)
        namecell = 'C' + str(row)
        classcell = 'D' + str(row)
        scheduledHoursCell = 'E' + str(row)
        actualHoursCell = 'F' + str(row)
        attendancePercentageCell = 'G' + str(row)
        endDateCell = "H" + str(row)
        notesCell = 'I' + str(row)
        currentStudent = allStudents[x]
        
        SAPColor = "00ADD8E6" #blue
        SAP1200Color = "0000FF00" #green
        yellowColor = "00FFFF00" #yellow
        lowAttendanceColor = "00CBC3E3" #light purple
        SAPTextColor = "000000BB" #dark blue
        VAStudentColor = "00FD5A87" #red
        LOAColor = "00838383" #gray
        nightStudentColor = "007FEAFD" #light blue
        darkBlueColor = "00277EFF" #dark blue
        darkRedColor = "00FF0000" #dark red
        
        dateTimeHour410 = datetime.timedelta(days=17, hours=2) #410 hours
        dateTimeHour460 = datetime.timedelta(days=19, hours=4) #460 hours
        dateTimeHour860 = datetime.timedelta(days=35, hours=20) #860 hours
        dateTimeHour900 = datetime.timedelta(days=37, hours=12) #900 hours
        dateTimeHour910 = datetime.timedelta(days=37, hours=22) #910 hours
        dateTimeHour1150 = datetime.timedelta(days=47, hours=22) #1150 hours
        dateTimeHour1210 = datetime.timedelta(days=50, hours=10) #1210 hours
        dateTimeHour1450 = datetime.timedelta(days=60, hours=10) #1450 hours
        dateTimeHour1510 = datetime.timedelta(days=62, hours=22) #1510 hours
        #TODO add 1500 hour students into timedelta
        
        #reapply black font value to notes cell
        sheet[notesCell].font = Font(color="00000000")#black
          
        sheet[numbercell] = row-3
        sheet[numbercell].border = thin_border
        
        sheet[idcell] = currentStudent.id
        sheet[idcell].border = thin_border
        
        sheet[namecell] = currentStudent.name
        sheet[namecell].border = thin_border
        
        sheet[classcell] = currentStudent.studentType + "/" + currentStudent.hourAmount
        sheet[classcell].border = thin_border
        
        sheet[scheduledHoursCell] = currentStudent.scheduledHours
        sheet[scheduledHoursCell].border = thin_border
        
        sheet[actualHoursCell] = currentStudent.actualHours
        sheet[actualHoursCell].border = thin_border
        
        sheet[attendancePercentageCell] = currentStudent.attendancePercentage
        sheet[attendancePercentageCell].border = thin_border
        
        sheet[notesCell] = currentStudent.notes + currentStudent.LOA
        sheet[notesCell].border = thin_border
        
        #datetime object m/d/y
        if type(currentStudent.endDate) == datetime.datetime:
            sheet[endDateCell] = currentStudent.endDate.strftime('%m-%d-%y')
        else:
            sheet[endDateCell] = currentStudent.endDate
        sheet[endDateCell].border = thin_border
        sheet[endDateCell].alignment = Alignment(horizontal="right")
        
        row += 1
        
        #color class cell yellow if part of 1200 hour program
        if currentStudent.VA == True:
            sheet[idcell].fill = PatternFill(start_color= VAStudentColor, end_color= VAStudentColor, fill_type= "solid")
            sheet[namecell].fill = PatternFill(start_color= VAStudentColor, end_color= VAStudentColor, fill_type= "solid")
            sheet[classcell].fill = PatternFill(start_color= VAStudentColor, end_color= VAStudentColor, fill_type= "solid")
            sheet[scheduledHoursCell].fill = PatternFill(start_color= VAStudentColor, end_color= VAStudentColor, fill_type= "solid")
            sheet[actualHoursCell].fill = PatternFill(start_color= VAStudentColor, end_color= VAStudentColor, fill_type= "solid")
        
        
        if "Night" in currentStudent.studentType:
            if isinstance(currentStudent.actualHours, datetime.time):
                sheet[classcell].fill = PatternFill(start_color= nightStudentColor, end_color= nightStudentColor, fill_type= "solid")
            elif currentStudent.actualHours > dateTimeHour900:
                sheet[classcell].fill = PatternFill(start_color= darkBlueColor, end_color= darkBlueColor, fill_type= "solid")
            else:
                sheet[classcell].fill = PatternFill(start_color= nightStudentColor, end_color= nightStudentColor, fill_type= "solid")
                
        if "Day" in currentStudent.studentType:
            if isinstance(currentStudent.actualHours, datetime.time):
                sheet[classcell].fill = PatternFill(start_color= yellowColor, end_color= yellowColor, fill_type= "solid")
            elif currentStudent.actualHours > dateTimeHour900:
                sheet[classcell].fill = PatternFill(start_color= darkRedColor, end_color= darkRedColor, fill_type= "solid")
            else:
                sheet[classcell].fill = PatternFill(start_color= yellowColor, end_color= yellowColor, fill_type= "solid")
            
        
        if "LOA" in currentStudent.LOA:
            sheet[idcell].fill = PatternFill(start_color = LOAColor, end_color=LOAColor, fill_type = "solid")
            sheet[namecell].fill = PatternFill(start_color = LOAColor, end_color=LOAColor, fill_type = "solid")
            sheet[classcell].fill = PatternFill(start_color = LOAColor, end_color=LOAColor, fill_type = "solid")
            sheet[scheduledHoursCell].fill = PatternFill(start_color = LOAColor, end_color=LOAColor,  fill_type = "solid")
            sheet[actualHoursCell].fill = PatternFill(start_color = LOAColor, end_color=LOAColor,  fill_type = "solid")
            sheet[attendancePercentageCell].fill = PatternFill(start_color = LOAColor, end_color=LOAColor,  fill_type = "solid")
            sheet[endDateCell].fill = PatternFill(start_color = LOAColor, end_color=LOAColor,  fill_type = "solid")
            sheet[notesCell].fill = PatternFill(start_color = LOAColor, end_color=LOAColor,  fill_type = "solid")
            
        if currentStudent.attendancePercentage <= 75:
            sheet[attendancePercentageCell].fill = PatternFill(start_color =lowAttendanceColor, end_color=lowAttendanceColor, fill_type = "solid")

        # Skips current loop if current hours are less than 24
        if type(currentStudent.actualHours) is datetime.time or type(currentStudent.actualHours) is datetime.datetime:
            continue
        
        if dateTimeHour410 <= currentStudent.actualHours <= dateTimeHour460:
            sheet[idcell].fill = PatternFill(start_color = SAPColor, end_color=SAPColor, fill_type = "solid")
            currentStudent.notes = "450 SAP; "
            
        if dateTimeHour860 <= currentStudent.actualHours <= dateTimeHour910:    
            sheet[idcell].fill = PatternFill(start_color = SAPColor, end_color=SAPColor, fill_type = "solid")
            currentStudent.notes = "900 SAP; " 
            
        if dateTimeHour1150 <= currentStudent.actualHours <= dateTimeHour1210:
            sheet[idcell].fill = PatternFill(start_color = SAP1200Color, end_color=SAP1200Color, fill_type = "solid")
            sheet[namecell].fill = PatternFill(start_color = SAP1200Color, end_color=SAP1200Color, fill_type = "solid")
            sheet[classcell].fill = PatternFill(start_color = SAP1200Color, end_color=SAP1200Color, fill_type = "solid")
            sheet[scheduledHoursCell].fill = PatternFill(start_color = SAP1200Color, end_color=SAP1200Color, fill_type = "solid")
            sheet[actualHoursCell].fill = PatternFill(start_color = SAP1200Color, end_color=SAP1200Color, fill_type = "solid")
            sheet[attendancePercentageCell].fill = PatternFill(start_color = SAP1200Color, end_color=SAP1200Color, fill_type = "solid")
            sheet[endDateCell].fill = PatternFill(start_color = SAP1200Color, end_color=SAP1200Color, fill_type = "solid")
            sheet[notesCell].fill = PatternFill(start_color = SAP1200Color, end_color=SAP1200Color, fill_type = "solid")
            currentStudent.notes = "1200 SAP; "    
            
        if dateTimeHour1450 <= currentStudent.actualHours <= dateTimeHour1510:
            sheet[idcell].fill = PatternFill(start_color = SAP1200Color, end_color=SAP1200Color, fill_type = "solid")
            sheet[namecell].fill = PatternFill(start_color = SAP1200Color, end_color=SAP1200Color, fill_type = "solid")
            sheet[classcell].fill = PatternFill(start_color = SAP1200Color, end_color=SAP1200Color, fill_type = "solid")
            sheet[scheduledHoursCell].fill = PatternFill(start_color = SAP1200Color, end_color=SAP1200Color, fill_type = "solid")
            sheet[actualHoursCell].fill = PatternFill(start_color = SAP1200Color, end_color=SAP1200Color, fill_type = "solid")
            sheet[attendancePercentageCell].fill = PatternFill(start_color = SAP1200Color, end_color=SAP1200Color, fill_type = "solid")
            sheet[endDateCell].fill = PatternFill(start_color = SAP1200Color, end_color=SAP1200Color, fill_type = "solid")
            sheet[notesCell].fill = PatternFill(start_color = SAP1200Color, end_color=SAP1200Color, fill_type = "solid")
            currentStudent.notes = "1500 SAP; "

        # Reapply the notes cell after checking for SAPs
        sheet[notesCell].font = Font(color=SAPTextColor)
        sheet[notesCell] = currentStudent.notes + currentStudent.LOA
        
    endingCell = 'G' + str(row)
    endingAverageCell = 'G' + str(row-1)
    totalEndingCell = 'F' + str(row)
    sheet[endingCell].value = '=AVERAGE(G4:' + str(endingAverageCell) + ')'
    sheet[totalEndingCell] = "Total:"
        
    
def countStudents():
    row = 1
    while True:
        cell = 'B' + str(row)
        if (sheet[cell].value != None):
            row += 1
        else:
            break
    return row - 1

#4513
def getStudentList():
    numStudents = countStudents()
    students = []
    row = 1
    for x in range(numStudents):
        idcell = 'B' + str(row)
        namecell = 'C' + str(row)
        scheduledHoursCell = 'K' + str(row)
        actualHoursCell = 'Q' + str(row)
        attendancePercentageCell = 'R' + str(row)
        newStudent = student((sheet[idcell].value),(sheet[namecell].value).upper(), (sheet[actualHoursCell].value), (sheet[scheduledHoursCell].value), (sheet[attendancePercentageCell].value))
        students.append(newStudent)
        row += 1
    return students

#6525
def getVAStatus(studentList):
    row = 1
    currentVACell = 'D' + str(row)
    currentIDCell = 'B' + str(row)
    while sheet[currentVACell].value != None:
        if "Yes" in sheet[currentVACell].value:
            for x in range(len(studentList)):
                if sheet[currentIDCell].value == studentList[x].id:
                    studentList[x].VA = True
                    break
        row += 1
        currentVACell = 'D' + str(row)
        currentIDCell = 'B' + str(row)

#4531
def addStudentType(studentList):
    sheet.delete_cols(16,1)
    sheet.delete_cols(2,13)
    row = 1
    currentIDCell = 'A' + str(row)
    currentClassCell = 'B' + str(row)
    while sheet[currentIDCell].value != None:
        if sheet[currentClassCell].value == 5 or sheet[currentClassCell].value.hour == 5:
            for x in range(len(studentList)):
                if sheet[currentIDCell].value == studentList[x].id:
                    studentList[x].studentType = "Night 20"
                    #print(studentList[x].name)
                    break
        if sheet[currentClassCell].value == 5.5 or (sheet[currentClassCell].value.hour == 5 and sheet[currentClassCell].value.minute == 30):
            for x in range(len(studentList)):
                if sheet[currentIDCell].value == studentList[x].id:
                    studentList[x].studentType = "Day 28"
                    #print(studentList[x].name)
                    break
        if sheet[currentClassCell].value == 6 or sheet[currentClassCell].value.hour == 6:
            for x in range(len(studentList)):
                if sheet[currentIDCell].value == studentList[x].id:
                    studentList[x].studentType = "Night 24"
                    #print(studentList[x].name)
                    break
        if sheet[currentClassCell].value == 7 or sheet[currentClassCell].value.hour == 7:
            for x in range(len(studentList)):
                if sheet[currentIDCell].value == studentList[x].id:
                    studentList[x].studentType = "Day 34"
                    #print(studentList[x].name)
                    break
        
        row += 1
        currentIDCell = 'A' + str(row)
        currentClassCell = 'B' + str(row)

#Sheet 1
def deleteHourReportColumns():
    sheet.delete_cols(21,1)
    sheet.delete_cols(14,5)
    sheet.delete_cols(4,9)
    sheet.delete_cols(1,1)
    
#4501
def addStudentListInfo(studentList):
    row = 1
    currentIDCell = 'A' + str(row)
    currentEndDateCell = 'L' + str(row)
    currentHourProgramCell = 'J' + str(row)
    currentLOACell = 'N' + str(row)
    while sheet[currentIDCell].value != None:
        for x in range(len(studentList)):
            if sheet[currentIDCell].value == studentList[x].id:
                studentList[x].endDate = sheet[currentEndDateCell].value
                if sheet[currentHourProgramCell].value == datetime.timedelta(days=41, hours=16):
                    studentList[x].hourAmount = "1000"
                if sheet[currentHourProgramCell].value == datetime.timedelta(days=50):
                    studentList[x].hourAmount = "1200"
                if sheet[currentHourProgramCell].value == datetime.timedelta(days=62, hours=12):
                    studentList[x].hourAmount = "1500"
                if type(sheet[currentEndDateCell].value) == datetime.datetime:
                    tempDate = sheet[currentEndDateCell].value
                    dateTimeConversionVar = datetime.date(year=tempDate.year, month=tempDate.month, day=tempDate.day)
                    studentList[x].endDate = dateTimeConversionVar
                if "LOA" in sheet[currentLOACell].value:
                    studentList[x].LOA = "LOA"
                break
        row += 1
        currentIDCell = 'A' + str(row)
        currentEndDateCell = 'L' + str(row)
        currentHourProgramCell = 'J' + str(row)
        currentLOACell = 'N' + str(row)
                    
    
def addReportFooter(row, studentList):
    currentTotalCell = 'B' + str(row)
    currentClassCell = 'C' + str(row)
    currentSAPCell = 'D' + str(row)
    currentDayInstructorCell = 'F' + str(row)
    currentNightInstructorCell = 'G' + str(row)
    currentColorCodeCell = 'I' + str(row)
    
    #Row 1
    sheet[currentTotalCell] = "Total"
    sheet[currentClassCell] = "Class"
    sheet[currentSAPCell] = "SAP"
    sheet.merge_cells('D' + str(row) + ":E" + str(row))
    sheet[currentDayInstructorCell] = "Day"
    sheet[currentNightInstructorCell] = "Night"
    #Row 2
    row += 1
    sheet[currentClassCell] = "Day 34"
    sheet[currentSAPCell] = "1)SAP  = 450 (1-450)"
    sheet.merge_cells('D' + str(row) + ":E" + str(row))
    sheet[currentDayInstructorCell] = "ANGIE"
    sheet[currentNightInstructorCell] = "KAT/JO"
    sheet[currentColorCodeCell] = "LOA"
    #Row 3
    row += 1
    sheet[currentClassCell] = "Day 28"
    sheet[currentSAPCell] = "2)SAP = 900 (451-900)"
    sheet.merge_cells('D' + str(row) + ":E" + str(row))
    sheet[currentDayInstructorCell] = "ANGIE"
    sheet[currentNightInstructorCell] = "KAT/JO"
    sheet[currentColorCodeCell] = "PRESTEN = add Tests start 3/6/23"
    #Row 4
    row += 1
    sheet[currentClassCell] = "Night 24"
    sheet[currentSAPCell] = "3)SAP = 1200 (901-1200)"
    sheet.merge_cells('D' + str(row) + ":E" + str(row))
    sheet[currentDayInstructorCell] = "NORMA"
    sheet[currentNightInstructorCell] = "LOU"
    sheet[currentColorCodeCell] = "RED COLOR = VA STUDENT"
    #Row 5
    row += 1
    sheet[currentClassCell] = "CROSSOVER/200"
    sheet[currentSAPCell] = "4)SAP  = 1500 (1201-1500)"
    sheet.merge_cells('D' + str(row) + ":E" + str(row))
    sheet[currentDayInstructorCell] = "NORMA"
    sheet[currentNightInstructorCell] = "LOU"
    sheet[currentColorCodeCell] = "GREEN = COMING SOON GRADS"
    

#Get all day students into one list
def getDayStudents(students):
    dayStudents = []
    for x in range(len(students)):
        if students[x].studentType == "Day 34" or students[x].studentType == "Day 28":
            dayStudents.append(students[x])
    return dayStudents

#Get all night students into one list
def getNightStudents(students):
    nightStudents = []
    for x in range(len(students)):
        if students[x].studentType == "Night 24" or students[x].studentType == "Night 20":
            nightStudents.append(students[x])
    return nightStudents

def get_dates_and_days_of_next_month():
    # Get the current date
    today = datetime.today()
    
    # Find the last day of the current month
    if today.month == 12:
        last_day_of_month = datetime(today.year + 1, 1, 1) - timedelta(days=1)
    else:
        last_day_of_month = datetime(today.year, today.month + 1, 1) - timedelta(days=1)
    
    # Check if today is within the last week of the month
    if today < last_day_of_month - timedelta(days=6):
        return None
    
    # Calculate the next month and year
    if today.month == 12:
        next_month = 1
        next_year = today.year + 1
    else:
        next_month = today.month + 1
        next_year = today.year
    
    # Create an empty list to store the results
    dates_and_days = []
    
    # Iterate through the days of the next month
    for day in range(1, 32): # 32 is used to cover all possible days in a month
        try:
            # Create a datetime object for the current day
            current_date = datetime(next_year, next_month, day)
            
            # Get the name of the day
            day_name = current_date.strftime("%A")
            
            # Append the date and day name to the result list
            dates_and_days.append((day, day_name))
        except ValueError:
            # Break the loop if the day is out of range for the given month
            break
            
    return dates_and_days

def fillSeparatedReport(allStudents, row):
    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
    
    for x in range(len(allStudents)):
        numbercell = 'A' + str(row)
        idcell = 'B' + str(row)
        namecell = 'C' + str(row)
        classcell = 'D' + str(row)
        scheduledHoursCell = 'E' + str(row)
        actualHoursCell = 'F' + str(row)
        attendancePercentageCell = 'G' + str(row)
        endDateCell = "H" + str(row)
        notesCell = 'I' + str(row)
        currentStudent = allStudents[x]
        
        SAPColor = "00ADD8E6" #blue
        SAP1200Color = "0000FF00" #green
        yellowColor = "00FFFF00" #yellow
        lowAttendanceColor = "00CBC3E3" #light purple
        SAPTextColor = "000000BB" #dark blue
        VAStudentColor = "00FD5A87" #red
        LOAColor = "00838383" #gray
        nightStudentColor = "007FEAFD" #light blue
        darkBlueColor = "00277EFF" #dark blue
        darkRedColor = "00FF0000" #dark red
        
        dateTimeHour410 = datetime.timedelta(days=17, hours=2) #410 hours
        dateTimeHour460 = datetime.timedelta(days=19, hours=4) #460 hours
        dateTimeHour860 = datetime.timedelta(days=35, hours=20) #860 hours
        dateTimeHour900 = datetime.timedelta(days=37, hours=12) #900 hours
        dateTimeHour910 = datetime.timedelta(days=37, hours=22) #910 hours
        dateTimeHour1150 = datetime.timedelta(days=47, hours=22) #1150 hours
        dateTimeHour1210 = datetime.timedelta(days=50, hours=10) #1210 hours
        dateTimeHour1450 = datetime.timedelta(days=60, hours=10) #1450 hours
        dateTimeHour1510 = datetime.timedelta(days=62, hours=22) #1510 hours
        #TODO add 1500 hour students into timedelta
        
        #reapply black font value to notes cell
        sheet[notesCell].font = Font(color="00000000")#black
          
        sheet[numbercell] = row
        sheet[numbercell].border = thin_border
        
        sheet[idcell] = currentStudent.id
        sheet[idcell].border = thin_border
        
        sheet[namecell] = currentStudent.name
        sheet[namecell].border = thin_border
        
        sheet[classcell] = currentStudent.studentType + "/" + currentStudent.hourAmount
        sheet[classcell].border = thin_border
        
        sheet[scheduledHoursCell] = currentStudent.scheduledHours
        sheet[scheduledHoursCell].border = thin_border
        
        sheet[actualHoursCell] = currentStudent.actualHours
        sheet[actualHoursCell].border = thin_border
        
        sheet[attendancePercentageCell] = currentStudent.attendancePercentage
        sheet[attendancePercentageCell].border = thin_border
        
        sheet[notesCell] = currentStudent.notes + currentStudent.LOA
        sheet[notesCell].border = thin_border
        
        #datetime object m/d/y
        if type(currentStudent.endDate) == datetime.datetime:
            sheet[endDateCell] = currentStudent.endDate.strftime('%m-%d-%y')
        else:
            sheet[endDateCell] = currentStudent.endDate
        sheet[endDateCell].border = thin_border
        sheet[endDateCell].alignment = Alignment(horizontal="right")
        
        row += 1
        
        #color class cell yellow if part of 1200 hour program
        if currentStudent.VA == True:
            sheet[idcell].fill = PatternFill(start_color= VAStudentColor, end_color= VAStudentColor, fill_type= "solid")
            sheet[namecell].fill = PatternFill(start_color= VAStudentColor, end_color= VAStudentColor, fill_type= "solid")
            sheet[namecell].font = Font(color="00FFFFFF")#white
            sheet[classcell].fill = PatternFill(start_color= VAStudentColor, end_color= VAStudentColor, fill_type= "solid")
            sheet[scheduledHoursCell].fill = PatternFill(start_color= VAStudentColor, end_color= VAStudentColor, fill_type= "solid")
            sheet[actualHoursCell].fill = PatternFill(start_color= VAStudentColor, end_color= VAStudentColor, fill_type= "solid")
        
        
        if "Night" in currentStudent.studentType:
            if isinstance(currentStudent.actualHours, datetime.time):
                sheet[classcell].fill = PatternFill(start_color= nightStudentColor, end_color= nightStudentColor, fill_type= "solid")
            elif currentStudent.actualHours > dateTimeHour900:
                sheet[classcell].fill = PatternFill(start_color= darkBlueColor, end_color= darkBlueColor, fill_type= "solid")
            else:
                sheet[classcell].fill = PatternFill(start_color= nightStudentColor, end_color= nightStudentColor, fill_type= "solid")
                
        if "Day" in currentStudent.studentType:
            if isinstance(currentStudent.actualHours, datetime.time):
                sheet[classcell].fill = PatternFill(start_color= yellowColor, end_color= yellowColor, fill_type= "solid")
            elif currentStudent.actualHours > dateTimeHour900:
                sheet[classcell].fill = PatternFill(start_color= darkRedColor, end_color= darkRedColor, fill_type= "solid")
                sheet[classcell].font = Font(color="00FFFFFF")#white
            else:
                sheet[classcell].fill = PatternFill(start_color= yellowColor, end_color= yellowColor, fill_type= "solid")
            
        
        if "LOA" in currentStudent.LOA:
            sheet[idcell].fill = PatternFill(start_color = LOAColor, end_color=LOAColor, fill_type = "solid")
            sheet[namecell].fill = PatternFill(start_color = LOAColor, end_color=LOAColor, fill_type = "solid")
            sheet[classcell].fill = PatternFill(start_color = LOAColor, end_color=LOAColor, fill_type = "solid")
            sheet[scheduledHoursCell].fill = PatternFill(start_color = LOAColor, end_color=LOAColor,  fill_type = "solid")
            sheet[actualHoursCell].fill = PatternFill(start_color = LOAColor, end_color=LOAColor,  fill_type = "solid")
            sheet[attendancePercentageCell].fill = PatternFill(start_color = LOAColor, end_color=LOAColor,  fill_type = "solid")
            sheet[endDateCell].fill = PatternFill(start_color = LOAColor, end_color=LOAColor,  fill_type = "solid")
            sheet[notesCell].fill = PatternFill(start_color = LOAColor, end_color=LOAColor,  fill_type = "solid")
            
        if currentStudent.attendancePercentage <= 75:
            sheet[attendancePercentageCell].fill = PatternFill(start_color =lowAttendanceColor, end_color=lowAttendanceColor, fill_type = "solid")

        # Skips current loop if current hours are less than 24
        if type(currentStudent.actualHours) is datetime.time or type(currentStudent.actualHours) is datetime.datetime:
            continue
        
        if dateTimeHour410 <= currentStudent.actualHours <= dateTimeHour460:
            sheet[idcell].fill = PatternFill(start_color = SAPColor, end_color=SAPColor, fill_type = "solid")
            currentStudent.notes = "450 SAP; "
            
        if dateTimeHour860 <= currentStudent.actualHours <= dateTimeHour910:    
            sheet[idcell].fill = PatternFill(start_color = SAPColor, end_color=SAPColor, fill_type = "solid")
            currentStudent.notes = "900 SAP; " 
            
        if dateTimeHour1150 <= currentStudent.actualHours <= dateTimeHour1210:
            sheet[idcell].fill = PatternFill(start_color = SAP1200Color, end_color=SAP1200Color, fill_type = "solid")
            sheet[namecell].fill = PatternFill(start_color = SAP1200Color, end_color=SAP1200Color, fill_type = "solid")
            sheet[classcell].fill = PatternFill(start_color = SAP1200Color, end_color=SAP1200Color, fill_type = "solid")
            sheet[scheduledHoursCell].fill = PatternFill(start_color = SAP1200Color, end_color=SAP1200Color, fill_type = "solid")
            sheet[actualHoursCell].fill = PatternFill(start_color = SAP1200Color, end_color=SAP1200Color, fill_type = "solid")
            sheet[attendancePercentageCell].fill = PatternFill(start_color = SAP1200Color, end_color=SAP1200Color, fill_type = "solid")
            sheet[endDateCell].fill = PatternFill(start_color = SAP1200Color, end_color=SAP1200Color, fill_type = "solid")
            sheet[notesCell].fill = PatternFill(start_color = SAP1200Color, end_color=SAP1200Color, fill_type = "solid")
            currentStudent.notes = "1200 SAP; "    
            
        if dateTimeHour1450 <= currentStudent.actualHours <= dateTimeHour1510:
            sheet[idcell].fill = PatternFill(start_color = SAP1200Color, end_color=SAP1200Color, fill_type = "solid")
            sheet[namecell].fill = PatternFill(start_color = SAP1200Color, end_color=SAP1200Color, fill_type = "solid")
            sheet[classcell].fill = PatternFill(start_color = SAP1200Color, end_color=SAP1200Color, fill_type = "solid")
            sheet[scheduledHoursCell].fill = PatternFill(start_color = SAP1200Color, end_color=SAP1200Color, fill_type = "solid")
            sheet[actualHoursCell].fill = PatternFill(start_color = SAP1200Color, end_color=SAP1200Color, fill_type = "solid")
            sheet[attendancePercentageCell].fill = PatternFill(start_color = SAP1200Color, end_color=SAP1200Color, fill_type = "solid")
            sheet[endDateCell].fill = PatternFill(start_color = SAP1200Color, end_color=SAP1200Color, fill_type = "solid")
            sheet[notesCell].fill = PatternFill(start_color = SAP1200Color, end_color=SAP1200Color, fill_type = "solid")
            currentStudent.notes = "1500 SAP; "

        # Reapply the notes cell after checking for SAPs
        sheet[notesCell].font = Font(color=SAPTextColor)
        sheet[notesCell] = currentStudent.notes + currentStudent.LOA

    return row
    
# Goes inside that folder. 

root = tkinter.Tk()
root.withdraw()

file_path = filedialog.askdirectory()
os.chdir(file_path)

wb = openpyxl.load_workbook("Book1.xlsx")

sheet = wb['4513']
students = getStudentList()
sheet = wb['4531']
addStudentType(students)
sheet = wb['4501']
addStudentListInfo(students)
sheet = wb['6525']
getVAStatus(students)
wb.save("test.xlsx")
    
wb.close()
#Create Report Template
wb = openpyxl.Workbook()
sheet = wb.active
createTemplate()
fillReport(students, 4)

#Logic for other report goes here (Separation by day/night + hour amount)
#Need to separate: 1-250, 251-450, 451-900, 901-1200 for day and night students (2 sheets total)
sheet = wb.create_sheet("Night Student Hour Separation")

nightStudents = getNightStudents(students)
#Separate students into different lists depending on their hour amount
nightStudents250 = []
nightStudents450 = []
nightStudents900 = []
nightStudents1000 = []
nightStudents1200 = []

for student in nightStudents:
    if type(student.actualHours) is datetime.time:
        nightStudents250.append(student)
        continue
    if student.actualHours <= datetime.timedelta(days=10, hours=10):
        nightStudents250.append(student)
    elif datetime.timedelta(days=10) < student.actualHours <= datetime.timedelta(days=18, hours=18):
        nightStudents450.append(student)
    elif datetime.timedelta(days=18, hours=18) < student.actualHours <= datetime.timedelta(days=37, hours=12):
        nightStudents900.append(student)
    elif datetime.timedelta(days=37, hours=12) < student.actualHours:
        nightStudents1200.append(student)

createTemplate()
titleCell = "C5"
teamCell = "I5"
sheet[titleCell] = "1-250 FRESHMEN (friend & family only no charge)"
sheet[teamCell] = "WHITE TEAM"
secondTeam = fillSeparatedReport(nightStudents250, 6)
titleCell = "C" + str(secondTeam+2)
teamCell = "I" + str(secondTeam+2)
sheet[titleCell] = "251-450 SOPHOMORE MAIN FLOOR"
sheet[teamCell] = "YELLOW TEAM"
thirdTeam = fillSeparatedReport(nightStudents450, secondTeam+3)
titleCell = "C" + str(thirdTeam+2)
teamCell = "I" + str(thirdTeam+2)
sheet[titleCell] = "451-900 SENIOR MAIN FLOOR (all services)"
sheet[teamCell] = "RED TEAM"
fourthTeam = fillSeparatedReport(nightStudents900, thirdTeam+3)
titleCell = "C" + str(fourthTeam+2)
teamCell = "I" + str(fourthTeam+2)
sheet[titleCell] = "901-1200/1500 SB (complete Tests & Time card)"
sheet[teamCell] = "GREEN TEAM"
finalRow = fillSeparatedReport(nightStudents1200, fourthTeam+3)
#nextRow = fillSeparatedReport(dayStudents, row)
#finishRow = fillSeparatedReport(nightStudents, nextRow+3)


sheet = wb.create_sheet("Day Student Hour Separation")

dayStudents = getDayStudents(students)
#Separate students into different lists depending on their hour amount
dayStudents250 = []
dayStudents450 = []
dayStudents900 = []
dayStudents1200 = []

for student in dayStudents:
    if type(student.actualHours) is datetime.time:
        dayStudents250.append(student)
        continue
    if student.actualHours <= datetime.timedelta(days=10, hours=10):
        dayStudents250.append(student)
    elif datetime.timedelta(days=10) < student.actualHours <= datetime.timedelta(days=18, hours=18):
        dayStudents450.append(student)
    elif datetime.timedelta(days=18, hours=18) < student.actualHours <= datetime.timedelta(days=37, hours=12):
        dayStudents900.append(student)
    elif datetime.timedelta(days=37, hours=12) < student.actualHours:
        dayStudents1200.append(student)

createTemplate()
titleCell = "C5"
teamCell = "I5"
sheet[titleCell] = "1-250 FRESHMEN (friend & family only no charge)"
sheet[teamCell] = "WHITE TEAM"
secondTeam = fillSeparatedReport(dayStudents250, 6)
titleCell = "C" + str(secondTeam+2)
teamCell = "I" + str(secondTeam+2)
sheet[titleCell] = "251-450 SOPHOMORE MAIN FLOOR"
sheet[teamCell] = "YELLOW TEAM"
thirdTeam = fillSeparatedReport(dayStudents450, secondTeam+3)
titleCell = "C" + str(thirdTeam+2)
teamCell = "I" + str(thirdTeam+2)
sheet[titleCell] = "451-900 SENIOR MAIN FLOOR (all services)"
sheet[teamCell] = "RED TEAM"
fourthTeam = fillSeparatedReport(dayStudents900, thirdTeam+3)
titleCell = "C" + str(fourthTeam+2)
teamCell = "I" + str(fourthTeam+2)
sheet[titleCell] = "901-1200/1500 SB (complete Tests & Time card)"
sheet[teamCell] = "GREEN TEAM"
finalRow = fillSeparatedReport(dayStudents1200, fourthTeam+3)
#nextRow = fillSeparatedReport(dayStudents, row)
#finishRow = fillSeparatedReport(nightStudents, nextRow+3)

wb.save("Attendance Report " + datetime.datetime.now().strftime('%Y-%m-%d') + ".xlsx")
wb.close()

