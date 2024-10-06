import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins, PrintOptions
from openpyxl.worksheet.pagebreak import Break
import math

# Function to generate a list of placeholder student names
def generate_student_names(prefix, count):
    return [f"{prefix} Student {i+1}" for i in range(count)]

def set_column_width_in_inches(worksheet, column, width_in_inches):
    # 1 inch ≈ 10.71 Excel column width units
    excel_width = width_in_inches * 10.5
    worksheet.column_dimensions[column].width = excel_width

# Generate test arrays of students that overflow to multiple pages
# Adjust the number of students to ensure multiple pages are created
day_students = generate_student_names("Day", 100)    # For example, 100 day students
night_students = generate_student_names("Night", 80) # For example, 80 night students

# Create a new workbook
wb = openpyxl.Workbook()

def create_sheet(wb, sheet_title, student_names, class_name):
    # Create a new sheet
    ws = wb.create_sheet(title=sheet_title)
    
    # Remove default sheet if it exists
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Desired row height in points
    desired_row_height_points = 25  # You can adjust this value as needed
    column_a_width = 2.5
    column_b_width = .75
    column_c_width = .75
    column_d_width = .75
    column_e_width = .75
    column_f_width = 2

    # Estimate the number of rows per page
    # Assume default row height is 15 points (approx 0.21 inches)
    # Printable height = page height - top margin - bottom margin
    # Page height for Letter size = 11 inches
    # Margins are set to 0.5 inches
    printable_height = 11 - 0.5 - 0.5  # 10 inches
    desired_row_height_in_inches = desired_row_height_points / 72
    rows_per_page = int(printable_height / desired_row_height_in_inches)
    printable_width = 8.5 - 0.5 - 0.5  # 7.5 inches
    # Subtract the width of the first column (Student Name) to estimate the number of columns that fit on a page
    columns_per_page = 6  # Assume 6 columns fit on a page
    #TODO - Set column widths from original report to fit on one page
    # Subtract header rows
    header_rows = 4  # We have 4 header rows
    content_rows_per_page = rows_per_page - header_rows
    if content_rows_per_page <= 0:
        content_rows_per_page = 1  # Avoid division by zero
    
    total_content_rows = len(student_names)
    total_pages = math.ceil(total_content_rows / content_rows_per_page)

    # Define fills
    fill_black = PatternFill("solid", fgColor="000000")  # Black background
    fill_yellow = PatternFill("solid", fgColor="FFFF00")  # Yellow background
    fill_blue = PatternFill("solid", fgColor="0000FF")    # Blue background

    # Write the header rows
    # Row 1: "CR'U INSTITUTE OF COSMETOLOGY AND BARBERING" centered with white text and black background
    ws.merge_cells('A1:F1')
    ws['A1'] = "CR'U INSTITUTE OF COSMETOLOGY AND BARBERING"
    ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")  # White text
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = desired_row_height_points
    # Apply black fill to cells A1:F1
    for col in range(1, 7):  # Columns A(1) to F(6)
        cell = ws.cell(row=1, column=col)
        cell.fill = fill_black

    # Row 2: "Page X/Y" centered, "BARBER" right-aligned
    ws.merge_cells('A2:E2')  # Merge A2:E2 for centered "Page X/Y"
    ws['A2'] = f"Page 1/{total_pages}"
    ws['A2'].font = Font(bold=True)
    ws['A2'].alignment = Alignment(horizontal='center')
    ws['F2'] = "BARBER"
    ws['F2'].alignment = Alignment(horizontal='right')
    ws['F2'].font = Font(bold=True)
    ws.row_dimensions[2].height = desired_row_height_points

    # Row 3: "DATE ________________ ATTENDANCE SHEET" left-aligned, "CLASS" right-aligned and highlighted
    ws['A3'] = "DATE ________________ ATTENDANCE SHEET"
    ws['A3'].alignment = Alignment(horizontal='left')
    ws['A3'].font = Font(bold=True)
    ws['F3'] = class_name.upper()
    ws['F3'].alignment = Alignment(horizontal='right')
    ws['F3'].font = Font(bold=True)
    # Apply highlight based on class_name
    if class_name.upper() == "DAY":
        ws['F3'].fill = fill_yellow
    elif class_name.upper() == "NIGHT":
        ws['F3'].fill = fill_blue
    ws.row_dimensions[3].height = desired_row_height_points

    # Row 4: Column headers
    columns = ["Student Name", "In", "Break", "Return", "Out", "Signature"]
    for col_num, column_title in enumerate(columns, start=1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = column_title
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        # Set column widths for better readability
        ws.column_dimensions[get_column_letter(col_num)].width = 20
    ws.row_dimensions[4].height = desired_row_height_points

    # Define a thin border style for cells
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Apply border to the column header row
    for col_num in range(1, len(columns) + 1):
        cell = ws.cell(row=4, column=col_num)
        cell.border = thin_border

    # Write the student names starting from row 5
    start_row = 5
    for idx, student_name in enumerate(student_names):
        row_num = start_row + idx
        # Write the student name
        name_cell = ws.cell(row=row_num, column=1)
        name_cell.value = student_name
        name_cell.border = thin_border
        name_cell.alignment = Alignment(horizontal='left')
        # Apply borders to the rest of the cells in the row
        for col_num in range(2, len(columns) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        # Set row height
        ws.row_dimensions[row_num].height = desired_row_height_points
        ws.column_dimensions['A'].width = column_a_width

        # Check if we need to insert a page break
        if (idx + 1) % content_rows_per_page == 0 and (idx + 1) != total_content_rows:
            # Insert a page break before the next row
            ws.row_breaks.append(Break(id=row_num))
            # Update the "Page X/Y" value in the header of the next page
            page_number = (idx + 1) // content_rows_per_page + 1
            # Update the headers for the next page
            next_header_row = row_num + 1

            # Shift all subsequent rows down by 4 to make room for headers
            ws.insert_rows(next_header_row, amount=4)

            # Row 1: "CR'U INSTITUTE OF COSMETOLOGY AND BARBERING"
            ws.merge_cells(start_row=next_header_row, start_column=1, end_row=next_header_row, end_column=6)
            ws.cell(row=next_header_row, column=1).value = "CR'U INSTITUTE OF COSMETOLOGY AND BARBERING"
            ws.cell(row=next_header_row, column=1).font = Font(size=14, bold=True, color="FFFFFF")
            ws.cell(row=next_header_row, column=1).alignment = Alignment(horizontal='center')
            ws.row_dimensions[next_header_row].height = desired_row_height_points
            # Apply black fill
            for col in range(1, 7):
                cell = ws.cell(row=next_header_row, column=col)
                cell.fill = fill_black

            # Row 2: "Page X/Y" and "BARBER"
            ws.merge_cells(start_row=next_header_row + 1, start_column=1, end_row=next_header_row + 1, end_column=5)
            ws.cell(row=next_header_row + 1, column=1).value = f"Page {page_number}/{total_pages}"
            ws.cell(row=next_header_row + 1, column=1).font = Font(bold=True)
            ws.cell(row=next_header_row + 1, column=1).alignment = Alignment(horizontal='center')

            ws.cell(row=next_header_row + 1, column=6).value = "BARBER"
            ws.cell(row=next_header_row + 1, column=6).alignment = Alignment(horizontal='right')
            ws.cell(row=next_header_row + 1, column=6).font = Font(bold=True)
            ws.row_dimensions[next_header_row + 1].height = desired_row_height_points

            # Row 3: "DATE..." and "CLASS"
            ws.cell(row=next_header_row + 2, column=1).value = "DATE ________________ ATTENDANCE SHEET"
            ws.cell(row=next_header_row + 2, column=1).alignment = Alignment(horizontal='left')
            ws.cell(row=next_header_row + 2, column=1).font = Font(bold=True)

            ws.cell(row=next_header_row + 2, column=6).value = class_name.upper()
            ws.cell(row=next_header_row + 2, column=6).alignment = Alignment(horizontal='right')
            ws.cell(row=next_header_row + 2, column=6).font = Font(bold=True)
            # Apply highlight based on class_name
            if class_name.upper() == "DAY":
                ws.cell(row=next_header_row + 2, column=6).fill = fill_yellow
            elif class_name.upper() == "NIGHT":
                ws.cell(row=next_header_row + 2, column=6).fill = fill_blue
            ws.row_dimensions[next_header_row + 2].height = desired_row_height_points

            # Row 4: Column headers
            for col_num, column_title in enumerate(columns, start=1):
                cell = ws.cell(row=next_header_row + 3, column=col_num)
                cell.value = column_title
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
            ws.row_dimensions[next_header_row + 3].height = desired_row_height_points

            # Adjust the row number after inserting rows
            start_row += 4
    
        # Set column widths
    set_column_width_in_inches(ws, 'A', column_a_width)
    set_column_width_in_inches(ws, 'B', column_b_width)
    set_column_width_in_inches(ws, 'C', column_c_width)
    set_column_width_in_inches(ws, 'D', column_d_width)
    set_column_width_in_inches(ws, 'E', column_e_width)
    set_column_width_in_inches(ws, 'F', column_f_width)

    # Set print titles to repeat header rows on each printed page
    #ws.print_title_rows = '1:4'  # Repeat the first four rows (headers)

    # Set page margins and print options for a printable format
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)
    ws.print_options = PrintOptions(gridLines=True)

    # Set the sheet to fit to page width
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = False  # Allow the height to be variable

    # Adjust the scaling to fit more rows per page if needed
    ws.page_setup.scale = 100  # You can adjust this value as needed

    # Set paper size to Letter (8.5 x 11 inches)
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER

    # Optionally, set orientation to landscape if you have many columns
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT

    # Define print area
    max_row = ws.max_row
    ws.print_area = f'A1:F{max_row}'

# Remove the default sheet if it exists
if 'Sheet' in wb.sheetnames:
    default_sheet = wb['Sheet']
    wb.remove(default_sheet)

# Create sheets for day and night students
create_sheet(wb, "Day Students", day_students, "DAY")
create_sheet(wb, "Night Students", night_students, "NIGHT")

# Save the workbook to a file
wb.save('student_sign_in_sheet.xlsx')